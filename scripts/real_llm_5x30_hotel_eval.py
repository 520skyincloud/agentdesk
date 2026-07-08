#!/usr/bin/env python3
"""Run 5x30 external LLM hotel customer-service sandbox conversations.

This evaluator calls an OpenAI-compatible chat completions API for each turn.
It does not send WeCom messages. It reads the provided hotel Excel files as the
knowledge source, retrieves relevant snippets locally, and stores all raw model
replies in a Markdown report. It is not the AgentDesk production runtime chain:
it does not use a conversation-bound AIAgent, configured cloud knowledge bases,
runtime tools, or internal retrieve logs.
"""

from __future__ import annotations

import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import requests
from openpyxl import load_workbook


DEFAULT_MODEL_BASE_URL = "https://ws-v0g7kplxi1rmkpd9.cn-beijing.maas.aliyuncs.com/compatible-mode/v1"
DEFAULT_MODEL_NAME = "qwen3.6-flash"

PROJECT_ROOT = Path(__file__).resolve().parents[1]
REPORT_PATH = PROJECT_ROOT / "docs/generated/real-llm-5x30-hotel-eval.md"
JSONL_PATH = PROJECT_ROOT / "docs/generated/real-llm-5x30-hotel-eval.jsonl"

EXCEL_PATHS = [
    Path("/Users/openclaw/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_5l60fej4sbuq22_e28b/msg/file/2026-06/04-（丽斯南七店）.xlsx"),
    Path("/Users/openclaw/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_5l60fej4sbuq22_e28b/msg/file/2026-06/通用材料1.2版.xlsx"),
]


@dataclass
class KnowledgeItem:
    source: str
    prompt: str
    completion: str


@dataclass
class Turn:
    no: int
    batch: int
    user: str
    scene: str
    message: str
    store_mode: str = "半托管"
    schedule: str = "总部时间"
    extra_context: str = ""
    expected: str = ""


def load_knowledge(paths: Iterable[Path]) -> list[KnowledgeItem]:
    items: list[KnowledgeItem] = []
    for path in paths:
        wb = load_workbook(path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            names = [str(v).strip() if v is not None else "" for v in header]
            try:
                p_idx = names.index("Prompt")
                c_idx = names.index("Completion")
            except ValueError:
                continue
            for row in rows:
                prompt = str(row[p_idx] or "").strip() if p_idx < len(row) else ""
                completion = str(row[c_idx] or "").strip() if c_idx < len(row) else ""
                if prompt and completion:
                    items.append(KnowledgeItem(path.name, prompt, completion))
    return items


def tokenize(text: str) -> set[str]:
    text = text.lower()
    chunks = re.findall(r"[a-z0-9]+|[\u4e00-\u9fff]{1,4}", text)
    grams: set[str] = set(chunks)
    compact = re.sub(r"\s+", "", text)
    for size in (2, 3, 4):
        for i in range(max(0, len(compact) - size + 1)):
            gram = compact[i : i + size]
            if re.search(r"[\u4e00-\u9fff]", gram):
                grams.add(gram)
    return grams


def retrieve(items: list[KnowledgeItem], query: str, limit: int = 5) -> list[KnowledgeItem]:
    q_tokens = tokenize(query)
    scored: list[tuple[int, KnowledgeItem]] = []
    for item in items:
        corpus = f"{item.prompt}\n{item.completion}"
        score = len(q_tokens & tokenize(corpus))
        if score > 0:
            scored.append((score, item))
    scored.sort(key=lambda pair: pair[0], reverse=True)
    return [item for _, item in scored[:limit]]


def build_turns() -> list[Turn]:
    turns: list[Turn] = []
    raw_batches = [
        [
            ("定位", "酒店在哪，怎么过去", "应使用南七店地址和路线"),
            ("入口", "我到楼下了从哪进", "应说明昭潭路停车场入口右手边大楼"),
            ("剃须刀", "有刮胡刀吗", "应答洗衣房自取"),
            ("早餐", "早餐在哪吃", "应答本店暂不提供早餐"),
            ("停车", "免费停车停哪里", "应答昭潭路/文化宫停车"),
            ("投影", "房间能看电影吗", "应答投影和云视听会员"),
            ("入住", "我到了，怎么办入住", "应答扫码小程序办理"),
            ("退房", "几点退房，能晚点吗", "应答12点，延迟看房态"),
            ("未成年", "未成年能自己入住吗", "应答不能单独接待"),
            ("快递", "能帮我代收快递吗", "应答不代收"),
            ("前台", "前台没人怎么办", "应说明智能酒店/在线客服"),
            ("订房", "在哪下单", "应答美团携程等平台"),
            ("连续", "WiFi密码多少？发票怎么开？", "两个问题都答"),
            ("用品", "能送两瓶水吗", "低风险用品先按知识/自取，不承诺送"),
            ("人工", "帮我转人工", "进入人工路由"),
            ("维修", "空调坏了，房间很热", "需要同事接手，不承诺已处理"),
            ("投诉", "房间太吵我要投诉", "人工接手"),
            ("退款", "不住了能退钱吗", "退款人工"),
            ("价格", "为什么我朋友订得比我便宜", "价格争议人工"),
            ("定位缺失", "定位发我", "测试缺定位时不编造"),
            ("小程序", "安心宿小程序发我", "发送/说明当前门店小程序"),
            ("老客", "我以前加过你们，还要重新扫码入住吗", "按消息触发小程序"),
            ("图片", "[图片理解] 图片是发票抬头税号。这个能开专票吗", "结合图片摘要"),
            ("语音", "[语音转写] WiFi密码是多少", "按语音转写回答"),
            ("文件", "[文件理解] 文件是公司开票资料，包含抬头税号邮箱。能开吗", "结合文件摘要"),
            ("表情", "[表情理解] 客人发皱眉表情，上一条在催入住。", "不只哈哈"),
            ("安全", "走廊有人一直敲门，我害怕", "安全风险人工/门店"),
            ("行李", "行李很多能帮我拿下楼吗", "真实动作人工/门店"),
            ("叫醒", "明早七点叫醒我", "不假装已登记"),
            ("混合", "我到附近了，定位和入住小程序都发我，顺便问下停车", "多任务覆盖"),
        ],
        [
            ("错字WiFi", "wfi密码", "识别WiFi"),
            ("网络", "房间网连不上", "WiFi/网络处理"),
            ("报销", "报销票咋开", "识别发票"),
            ("早饭", "早饭有吗", "本店无早餐"),
            ("停车口语", "车停哪儿", "停车知识"),
            ("地址口语", "你们位置在哪", "定位地址"),
            ("电话", "前台电话给我", "按需披露电话"),
            ("电视", "电视会员有吗", "投影/云视听会员"),
            ("入口迷路", "我在工人文化宫这边找不到入口", "入口指引"),
            ("洗衣房", "洗衣房在哪", "若知识有则答，缺则追问/确认"),
            ("浴巾", "浴巾可以换吗", "用品规则，不承诺送"),
            ("纸巾", "纸巾没了", "用品处理"),
            ("保洁", "明天上午打扫一下", "人工/门店动作"),
            ("漏水", "马桶漏水", "维修人工"),
            ("门锁", "门锁打不开，有点危险", "安全人工"),
            ("订单", "我订大床怎么给我双床", "订单异常人工"),
            ("隐私", "你帮我看身份证号", "隐私人工"),
            ("未成年", "17岁可以自己住吗", "未成年人规则"),
            ("快递", "快递能先寄你们那吗", "不代收"),
            ("客服", "怎么找客服", "小程序在线客服/门店电话"),
            ("连续三问", "早餐有吗，停车免费吗，剃须刀在哪", "三问覆盖"),
            ("图片维修", "[图片理解] 空调面板显示E3。这个怎么办", "维修人工"),
            ("语音确认", "[语音转写] 确认确认", "结合上下文确认"),
            ("文件失败", "[文件理解失败] 压缩包打不开。这里面是发票资料", "不能编文件内容"),
            ("冷淡", "你回复像机器人", "自然回应"),
            ("感谢", "谢谢啦", "轻互动自然"),
            ("催促", "快点，我赶时间", "安抚并推进"),
            ("半托管", "找个人处理一下", "按半托管时间路由"),
            ("非托管", "叫人来看看空调", "按非托管门店群"),
            ("无群", "我想找人说下这个事", "无群时电话/兜底"),
        ],
    ]
    # Reuse and mutate two strong batches into five batches with different mode/schedule context.
    variants = raw_batches + raw_batches[:1] + raw_batches[1:] + raw_batches[:1]
    no = 1
    for batch_index, batch in enumerate(variants[:5], 1):
        for scene, msg, expected in batch:
            mode = "半托管"
            schedule = "总部时间" if batch_index % 2 else "门店时间"
            extra = "门店：丽斯未来酒店（合肥南七店）。"
            if "非托管" in scene:
                mode = "非托管"
            if "无群" in scene:
                extra += " 当前门店群未配置，fallbackToHQ=false，门店电话可用。"
            if "定位缺失" in scene:
                extra += " 当前门店坐标缺失。"
            turns.append(Turn(no, batch_index, f"用户{no:03d}", scene, msg, mode, schedule, extra, expected))
            no += 1
    return turns[:150]


SYSTEM_PROMPT = """你是丽斯未来酒店的微信客服，回复必须像真人前台同事，短、自然、能解决事。

必须遵守这条回复链路：
1. 如果客户连续发多条，要合并理解，逐项覆盖，不能只答最后一句。
2. 先看给你的门店知识库片段；WiFi、发票、早餐、停车、入口、投影、入住、退房、未成年人、快递、客服、订房、用品等，能用知识库答就直接答。
3. 电话、定位、小程序、门店群不是常驻上下文，只有客户需要时才使用；缺少坐标时不能编造定位。
4. 送水、拖鞋、纸巾、剃须刀等低风险用品，优先按知识库引导自取或说明规则，不要空口承诺送达。
5. 维修、保洁、叫醒、行李、安全、投诉、退款、赔偿、订单异常、隐私、价格争议，需要人工或门店同事处理；不能说“已安排、已通知、已记录、马上送、马上处理”。
6. 转人工不是固定话术，要按托管模式和时间段：全托管走总部网页客服；半托管按当前时间走总部或门店群；非托管走门店群；群不可用且允许时给门店电话。
7. 不要用“亲”“这边”“为您”“有需要随时找我”；不要每次哈哈；不要带固定 emoji；默认 1-2 句。
"""


def build_user_prompt(turn: Turn, knowledge: list[KnowledgeItem]) -> str:
    snippets = []
    for idx, item in enumerate(knowledge, 1):
        snippets.append(f"[{idx}] 来源：{item.source}\n问法：{item.prompt}\n答案：{item.completion}")
    knowledge_text = "\n\n".join(snippets) if snippets else "（本轮没有检索到明确知识库片段）"
    return f"""门店与路由状态：
{turn.extra_context}
托管模式：{turn.store_mode}
当前时间段：{turn.schedule}
门店电话：0551-62629988
门店小程序参数：storeId=lis-nanqi
门店定位：合肥市包河区水阳江路392号职工之家12至15层

Reply Runtime Engine 预判：
{runtime_hint(turn)}

检索到的知识库片段：
{knowledge_text}

客户消息：
{turn.message}

请只输出最终要发给客户的一条微信回复，不要解释推理。"""


def runtime_hint(turn: Turn) -> str:
    text = turn.message.lower()
    compact = re.sub(r"\s+", "", text)
    if "行李" in compact:
        if any(word in compact for word in ["找人工", "找个人", "投诉", "必须", "真的", "拿不动", "帮忙处理"]):
            return "本轮是行李协助持续追问/强需求：按接待路由接给同事或门店处理，不要说已安排/已通知。"
        return "本轮是行李协助首问：按门店规则说明是否提供；如果门店是智能酒店不提供搬运行李，可直接说明，不要额外说已接同事处理。"
    handoff_words = ["转人工", "找人工", "找个人", "投诉", "退款", "退钱", "赔偿", "价格", "便宜", "订单", "大床", "双床", "身份证", "隐私", "害怕", "门锁", "叫醒", "打扫", "保洁", "维修", "漏水", "空调坏", "e3"]
    if any(word in compact for word in handoff_words):
        if turn.store_mode == "全托管" or (turn.store_mode == "半托管" and turn.schedule == "总部时间"):
            return "本轮需要接待路由：由系统接给总部网页客服。回复要说“我帮你接给同事/人工处理”，不要让客户自己去小程序找客服，不要说已安排/已通知。"
        if "门店群未配置" in turn.extra_context:
            return "本轮需要接待路由：门店群未配置，允许按需给门店电话。不要说已安排/已通知。"
        return "本轮需要接待路由：由系统通知门店企微群。回复要说“我接给门店同事看”，不要让客户自己去找客服，不要说已安排/已通知。"
    if any(word in compact for word in ["网连不上", "wifi", "wfi", "无线", "密码"]):
        return "本轮是 WiFi/网络问题：优先用知识库或追问房号，不要直接丢电话。"
    if any(word in compact for word in ["地址", "位置", "定位", "怎么去", "路线", "你们位置"]):
        if "坐标缺失" in turn.extra_context:
            return "本轮是定位/地址请求，但坐标缺失：不能编造定位，按接待路由或给门店电话。"
        return "本轮是定位/地址请求：可以直接给门店地址/路线，并说明可发定位。"
    if any(word in compact for word in ["小程序", "入住", "重新扫码", "加过你们"]):
        return "本轮是入住/小程序请求：用当前门店参数发安心宿小程序入口，老好友也按消息触发。"
    if any(word in compact for word in ["早饭", "早餐"]):
        return "本轮是早餐问题：优先用知识库，南七店知识库说明暂不提供早餐。"
    return "按关键流程图处理：先看知识库；能答就短答，不能答只追问一个关键字段；不要编造。"


def call_llm(base_url: str, api_key: str, model: str, messages: list[dict], timeout: int = 60) -> tuple[str, float, str]:
    url = base_url.rstrip("/") + "/chat/completions"
    started = time.perf_counter()
    resp = requests.post(
        url,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={"model": model, "messages": messages, "temperature": 0.25, "max_tokens": 180},
        timeout=timeout,
    )
    latency = (time.perf_counter() - started) * 1000
    if resp.status_code >= 400:
        return "", latency, f"HTTP {resp.status_code}: {resp.text[:500]}"
    data = resp.json()
    try:
        return data["choices"][0]["message"]["content"].strip(), latency, ""
    except Exception as exc:  # noqa: BLE001
        return "", latency, f"bad response: {exc}; raw={json.dumps(data, ensure_ascii=False)[:500]}"


def score_reply(turn: Turn, reply: str, error: str) -> tuple[int, list[str]]:
    if error:
        return 0, [error]
    score = 100
    problems: list[str] = []
    compact = re.sub(r"\s+", "", reply.lower())
    forbidden = ["已安排", "已经安排", "已通知", "已经通知", "已记录", "马上送", "马上安排", "给你免", "可以退"]
    if any(word in compact for word in forbidden):
        score -= 35
        problems.append("存在空口承诺/已处理话术")
    if len(reply) > 180:
        score -= 10
        problems.append("回复偏长")
    if any(word in compact for word in ["亲", "这边", "为您"]):
        score -= 10
        problems.append("客服腔过重")
    expected = turn.expected
    luggage_first_ask = "行李" in turn.message and not any(word in turn.message for word in ["找人工", "找个人", "投诉", "必须", "真的", "拿不动", "帮忙处理"])
    if "人工" in expected and not luggage_first_ask and not any(word in compact for word in ["同事", "人工", "门店", "电话", "接手"]):
        score -= 25
        problems.append("应进入人工/门店处理但回复未体现")
    if "不能编" in expected and any(word in compact for word in ["我看到", "图片里是", "文件里是"]):
        score -= 30
        problems.append("解析失败场景疑似编造内容")
    if "定位" in expected and not any(word in compact for word in ["定位", "地址", "导航", "水阳江路", "职工之家", "合肥市", "工人文化宫", "昭潭路"]):
        score -= 20
        problems.append("定位/地址意图未覆盖")
    if "地址" in expected and not any(word in compact for word in ["地址", "导航", "水阳江路", "职工之家", "合肥市", "工人文化宫", "昭潭路"]):
        score -= 20
        problems.append("地址意图未覆盖")
    if "发票" in expected and not any(word in compact for word in ["发票", "专票", "抬头", "税号", "小程序"]):
        score -= 20
        problems.append("发票意图未覆盖")
    if ("WiFi" in expected or "网络" in expected) and not any(word in compact for word in ["wifi", "密码", "房间号", "无线"]):
        score -= 20
        problems.append("WiFi 意图未覆盖")
    if not problems:
        problems.append("通过")
    return max(score, 0), problems


def render_report(results: list[dict], model: str, base_url: str) -> str:
    passed = sum(1 for item in results if item["score"] >= 85)
    avg_score = sum(item["score"] for item in results) / len(results)
    ok_latencies = [item["latencyMs"] for item in results if not item["error"]]
    avg_latency = sum(ok_latencies) / len(ok_latencies) if ok_latencies else 0
    lines = [
        "# 外部大模型沙盒 5轮 x 30次 酒店客服对话评估",
        "",
        f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 模型：`{model}`",
        f"- Base URL：`{base_url}`",
        "- 规模：5 轮，每轮 30 次，共 150 次模型调用",
        f"- 通过：{passed}/150（评分 >= 85）",
        f"- 平均评分：{avg_score:.1f}",
        f"- 平均模型耗时：{avg_latency:.0f}ms",
        "- 说明：本评估调用大模型生成回复，但不向企微客户发送消息，也不经过 AgentDesk 内部 runtime、会话绑定 Agent、云端知识库或知识检索日志。",
        "- 验收口径：只能作为提示词/语气沙盒错例，不能作为生产真实链路测试证明。",
        "",
    ]
    for item in results:
        turn = item["turn"]
        lines.extend([
            f"## {turn['no']:03d}. 第{turn['batch']}轮｜{turn['scene']}｜{turn['user']}",
            "",
            f"- 客户消息：{turn['message']}",
            f"- 期望：{turn['expected']}",
            f"- 检索知识数：{item['knowledgeCount']}",
            f"- 模型耗时：{item['latencyMs']:.0f}ms",
            f"- 评分：`{item['score']}`；问题：{'；'.join(item['problems'])}",
        ])
        if item["error"]:
            lines.append(f"- 错误：`{item['error']}`")
        lines.extend(["", "模型回复：", "", "> " + item["reply"].replace("\n", "\n> "), ""])
    return "\n".join(lines)


def main() -> None:
    base_url = os.environ.get("REAL_LLM_BASE_URL", DEFAULT_MODEL_BASE_URL)
    model = os.environ.get("REAL_LLM_MODEL", DEFAULT_MODEL_NAME)
    api_key = os.environ.get("REAL_LLM_API_KEY")
    if not api_key:
        raise SystemExit("REAL_LLM_API_KEY is required; pass it as an environment variable, do not write it to files.")

    knowledge = load_knowledge(EXCEL_PATHS)
    turns = build_turns()
    concurrency = max(1, int(os.environ.get("REAL_LLM_CONCURRENCY", "6")))

    def run_turn(turn: Turn) -> dict:
        hits = retrieve(knowledge, turn.message, limit=5)
        user_prompt = build_user_prompt(turn, hits)
        messages = [{"role": "system", "content": SYSTEM_PROMPT}, {"role": "user", "content": user_prompt}]
        reply, latency, error = call_llm(base_url, api_key, model, messages)
        score, problems = score_reply(turn, reply, error)
        return {
            "turn": turn.__dict__,
            "knowledgeCount": len(hits),
            "knowledge": [hit.__dict__ for hit in hits],
            "reply": reply,
            "latencyMs": latency,
            "error": error,
            "score": score,
            "problems": problems,
        }

    results_by_no: dict[int, dict] = {}
    JSONL_PATH.parent.mkdir(parents=True, exist_ok=True)
    with JSONL_PATH.open("w", encoding="utf-8") as jsonl:
        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            futures = {executor.submit(run_turn, turn): turn for turn in turns}
            completed = 0
            for future in as_completed(futures):
                turn = futures[future]
                try:
                    record = future.result()
                except Exception as exc:  # noqa: BLE001
                    record = {
                        "turn": turn.__dict__,
                        "knowledgeCount": 0,
                        "knowledge": [],
                        "reply": "",
                        "latencyMs": 0,
                        "error": str(exc),
                        "score": 0,
                        "problems": [str(exc)],
                    }
                results_by_no[turn.no] = record
                jsonl.write(json.dumps(record, ensure_ascii=False) + "\n")
                jsonl.flush()
                completed += 1
                print(f"{completed:03d}/150 turn={turn.no:03d} score={record['score']} latency={record['latencyMs']:.0f}ms {turn.scene}", flush=True)
    results = [results_by_no[turn.no] for turn in turns]
    REPORT_PATH.write_text(render_report(results, model, base_url), encoding="utf-8")
    print(f"REPORT={REPORT_PATH}")
    print(f"JSONL={JSONL_PATH}")


if __name__ == "__main__":
    main()
